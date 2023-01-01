from aiogram import Bot, types
from aiogram.dispatcher import Dispatcher
from aiogram.utils import executor
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, KeyboardButtonPollType
import asyncio
import logging
from openpyxl import load_workbook
import random
import sqlite3


ENC = 'utf-8'
FORMAT_INFO = '%(asctime)s, %(module)s, %(funcName)s, %(name)s, --> %(message)s'

logging.basicConfig(format=FORMAT_INFO, filename='project_1/log.txt',
                    filemode='a', encoding=ENC, level=logging.INFO, datefmt='%m/%d/%Y %H:%M:%S')

path = 'project_1/token_bot.config'  # файл с Токеном
d = open(path, 'r', encoding=ENC)
key = d.readline()
bot = Bot(token=key)
dp = Dispatcher(bot)

f_csv = 'project_1/phonebook_ooo.csv'
wb = load_workbook('project_1/phonebook_ooo.xlsx')
sheets = wb.sheetnames
sheet = wb[sheets[0]]

conn = sqlite3.connect('project_1/Users.db')
cur = conn.cursor()
with conn:
    data = conn.execute(
        "select count(*) from sqlite_master where type='table' and name='users'")
    for row in data:
        if row[0] == 0:
            with conn:
                conn.execute(
                    """CREATE TABLE users (chat_id VARCHAR(20) PRIMARY KEY,first_name VARCHAR(30),username VARCHAR(30));""")


@dp.message_handler(commands='start')
@dp.message_handler(text='Главная')
async def cmd_start(message: types.Message):
    logging.info(message)
    if not cur.execute(f'''select chat_id From users
                        where chat_id = {message.chat.id} ''').fetchall():
        cur.execute("INSERT INTO users(chat_id, first_name, username)"
                    "VALUES(?, ?, ?)",
                    (message.chat.id, message.chat.first_name, message.chat.username))
        conn.commit()
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    buttons = ['Телефонный справочник', 'Калькулятор']
    keyboard.add(*buttons)
    Game = types.KeyboardButton('Игра в монетки')
    keyboard.row(Game)
    await bot.send_message(chat_id=message.chat.id, text=(f'Привет {message.chat.first_name}. Что запустить?'), reply_markup=keyboard)


@dp.message_handler(text='Телефонный справочник')
async def phone(message: types.Message):
    logging.info(message.text)
    home = types.KeyboardButton('Главная')
    management = types.KeyboardButton('Руководство')
    administration = types.KeyboardButton('Аппарат при руководстве')
    department_1 = types.KeyboardButton('Отдел 1')
    department_2 = types.KeyboardButton('Отдел 2')
    department_3 = types.KeyboardButton('Отдел 3')
    department_4 = types.KeyboardButton('Отдел 4')
    department_5 = types.KeyboardButton('Отдел 5')
    department_6 = types.KeyboardButton('Отдел 6')
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.row(home).row(management, administration).row(department_1, department_2, department_3)\
            .row(department_4, department_5, department_6)
    await bot.send_message(chat_id=message.chat.id, text='Запускаю телефонный справочник\
    \nДля просмотра служебных телефонов работников ООО "Рога и копыта"\
    нажмите на иконку интересующего вас подразделения', reply_markup=keyboard)


@dp.message_handler(text=['Руководство', 'Аппарат при руководстве', 'Отдел 1',
                          'Отдел 2', 'Отдел 3', 'Отдел 4', 'Отдел 5', 'Отдел 6'])
async def select(message: types.Message):
    logging.info(message.text)
    data = sheet.rows
    csv = open(f_csv, 'w', encoding=ENC)
    for row in data:
        l = list(row)
        for i in range(len(l)):
            if i == len(l) - 1:
                csv.write(str(l[i].value)+'\n')
            else:
                csv.write(str(l[i].value) + ',')
    csv.close()
    with open('project_1/phonebook_ooo.csv', 'r', encoding=ENC) as d:
        while True:
            cont = d.readline()
            if not cont:
                break
            if message.text in cont:
                cont = cont.split(',')
                cont = list(map(str, cont))
                await bot.send_message(chat_id=message.chat.id, text=(f'{cont[5]} {cont[0]}\nТел.(внутренний) - {cont[6]}\nТел.(город) - {cont[7]}\
                \nТел.(сот) - {cont[8]}\nE-mail - {cont[9]}'))
    csv.close()


@dp.message_handler(text='Калькулятор')
async def inp(message: types.Message):
    logging.info(message.text)
    zer = types.KeyboardButton('0')
    one = types.KeyboardButton('1')
    two = types.KeyboardButton('2')
    three = types.KeyboardButton('3')
    four = types.KeyboardButton('4')
    five = types.KeyboardButton('5')
    six = types.KeyboardButton('6')
    seven = types.KeyboardButton('7')
    eight = types.KeyboardButton('8')
    nine = types.KeyboardButton('9')
    home = types.KeyboardButton('Главная')
    mult = types.KeyboardButton('*')
    div = types.KeyboardButton('/')
    summ = types.KeyboardButton('+')
    diff = types.KeyboardButton('-')
    res = types.KeyboardButton('=')
    sep = types.KeyboardButton('.')
    a = types.KeyboardButton('(')
    b = types.KeyboardButton(')')
    delit = types.KeyboardButton('del')
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.row(seven, eight, nine, mult, home).row(four, five, six, div, delit).row(one, two, three, diff, summ)\
            .row(zer, a, b, sep, res)
    await bot.send_message(chat_id=message.chat.id, text='Режим калькулятора', reply_markup=keyboard)
    mes = []

    @dp.message_handler(text=['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '.', '+', '-', '/', '*', 'del', '(', ')', '='])
    async def calc(message: types.Message):
        if message.text != '=':
            if message.text != 'del':
                mes.append(message.text)
            else:
                for _ in range(0, 1):
                    mes.pop()
                    await bot.send_message(chat_id=message.chat.id, text=(''.join(mes)))
        else:
            await message.answer(text=(''.join(mes)))
            res = eval(''.join(mes))
            await bot.send_message(chat_id=message.chat.id, text=res)
            logging.info(f'{mes} = {res}')
            mes.clear()


@dp.message_handler(text='Игра в монетки')
async def note(message: types.Message):
    opp = int(random.randint(1, 7))
    arr_pl = []
    arr_bot = []
    rez_pl, rez_bot, j = 0, 0, 0
    await message.answer(text='Предлагаю сыграть со мной в игру на доверие. Правила игры просты:\nЕсть некий автомат в \
который два игрока по очереди могут опустить (Доверие) \
или не опускать (Обман) монетку.\n- Если оба игрока доверяют друг другу (оба опустили в автомат по монетке) то каждый из них\
получает в награду по 2 монеты.\n- Если оба игрока обманули друг друга (ни кто из обоих игроков не опустил в автомат монетку),\
то ни один из игроков награды не получает.\n- Если один из игроков обманывает, а второй доверяет, то обманщик получает в \
награду три монеты, доверившийся игрок награды не получает.\
\nПобеждает тот игрок, у которго по итогу 10-ти ходов больше монет.')
    trust = types.KeyboardButton('Опустить монетку')
    cheat = types.KeyboardButton('Обмануть')
    home = types.KeyboardButton('Главная')
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.row(trust, cheat, home)
    await bot.send_message(chat_id=message.chat.id, text='Начнем? Делайте ход.', reply_markup=keyboard)

    @dp.message_handler(text=['Опустить монетку', 'Обмануть'])
    async def game_start(message: types.Message,):
        nonlocal j, rez_bot, rez_pl
        logging.info(message.text)
        match (message.text):
            case 'Обмануть':
                arr_pl.append(0)
            case 'Опустить монетку':
                arr_pl.append(1)
        logging.info(arr_pl)
        await bot.send_message(chat_id=message.chat.id, text=(f'Ход: {j+1}'))
        match (opp):
            case 1:  # Наивный (всегда доверяет)
                arr_bot.append(1)
                if arr_pl[j] == 1:
                    rez_pl = rez_pl + 2
                    rez_bot = rez_bot + 2
                    j += 1
                    await true_true(rez_pl, rez_bot)
                else:
                    rez_pl = rez_pl + 3
                    j += 1
                    await false_true(rez_pl, rez_bot)
            case 2:  # Обманщик (всегда обманывает)
                if arr_pl[j] == 1:
                    rez_bot = rez_bot + 3
                    await true_false(rez_pl, rez_bot)
                    arr_bot.append(0)
                    j += 1
                else:
                    await false_false(rez_pl, rez_bot)
                    arr_bot.append(0)
                    j += 1
            case 3:  # Подражатель (начинает с доверия далее повторяет ваш предыдущий ход)
                if j == 0:
                    if arr_pl[j] == 1:
                        rez_pl = rez_pl + 2
                        rez_bot = rez_bot + 2
                        await true_true(rez_pl, rez_bot)
                        arr_bot.append(1)
                        j += 1
                    else:
                        rez_pl = rez_pl + 3
                        await false_true(rez_pl, rez_bot)
                        arr_bot.append(1)
                        j += 1
                else:
                    if arr_pl[j] == 1 and arr_pl[j-1] == 1:
                        rez_pl = rez_pl + 2
                        rez_bot = rez_bot + 2
                        await true_true(rez_pl, rez_bot)
                        arr_bot.append(1)
                        j += 1
                    elif arr_pl[j] == 1 and arr_pl[j-1] == 0:
                        rez_bot = rez_bot + 3
                        await true_false(rez_pl, rez_bot)
                        arr_bot.append(0)
                        j += 1
                    elif arr_pl[j] == 0 and arr_pl[j-1] == 1:
                        rez_pl = rez_pl + 3
                        await false_true(rez_pl, rez_bot)
                        arr_bot.append(1)
                        j += 1
                    elif arr_pl[j] == 0 and arr_pl[j-1] == 0:
                        await false_false(rez_pl, rez_bot)
                        arr_bot.append(0)
                        j += 1
            # Злопамятный (начинает с доверия, если хоть раз был обманут обманывает до конца)
            case 4:
                if arr_pl[j] == 1 and 0 not in arr_pl:
                    rez_pl = rez_pl + 2
                    rez_bot = rez_bot + 2
                    await true_true(rez_pl, rez_bot)
                    arr_bot.append(1)
                    j += 1
                elif arr_pl[j] == 0 and sum(arr_pl) == len(arr_pl)-1:
                    rez_pl = rez_pl + 3
                    await false_true(rez_pl, rez_bot)
                    arr_bot.append(1)
                    j += 1
                elif arr_pl[j] == 1 and sum(arr_pl) < len(arr_pl):
                    rez_bot = rez_bot + 3
                    await true_false(rez_pl, rez_bot)
                    arr_bot.append(0)
                    j += 1
                elif arr_pl[j] == 0 and sum(arr_pl) < len(arr_pl) - 1:
                    await false_false(rez_pl, rez_bot)
                    arr_bot.append(0)
                    j += 1
            case 5:  # Детектив (1.Доверие, 2.Обман, 3. Доверие, 4. Доверие если в ответ на 2 ход обман - далее стратегия подражателя, если нет - Обманщик)
                match(j):
                    case 0:  # Ход 1
                        if arr_pl[j] == 0:
                            rez_pl = rez_pl + 3
                            await false_true(rez_pl, rez_bot)
                            arr_bot.append(1)
                            j += 1
                        else:
                            rez_pl = rez_pl + 2
                            rez_bot = rez_bot + 2
                            await true_true(rez_pl, rez_bot)
                            arr_bot.append(1)
                            j += 1
                    case 1:  # Ход 2
                        if arr_pl[j] == 0:
                            await false_false(rez_pl, rez_bot)
                            arr_bot.append(0)
                            j += 1
                        else:
                            rez_bot = rez_bot + 3
                            await true_false(rez_pl, rez_bot)
                            arr_bot.append(0)
                            j += 1
                    case 2:  # Ход 3
                        if arr_pl[j] == 0:
                            rez_pl = rez_pl + 3
                            await false_true(rez_pl, rez_bot)
                            arr_bot.append(1)
                            j += 1
                        else:
                            rez_pl = rez_pl + 2
                            rez_bot = rez_bot + 2
                            await true_true(rez_pl, rez_bot)
                            arr_bot.append(1)
                            j += 1
                    case 3:  # Ход 4
                        if arr_pl[j] == 0:
                            rez_pl = rez_pl + 3
                            await false_true(rez_pl, rez_bot)
                            arr_bot.append(1)
                            j += 1
                        else:
                            rez_pl = rez_pl + 2
                            rez_bot = rez_bot + 2
                            await true_true(rez_pl, rez_bot)
                            arr_bot.append(1)
                            j += 1
                    case _:  # Ход 5 и далее
                        if arr_pl[2] == 0:  # Выбрана стртегия Подражатель
                            if arr_pl[j] == 1 and arr_pl[j-1] == 1:
                                rez_pl = rez_pl + 2
                                rez_bot = rez_bot + 2
                                await true_true(rez_pl, rez_bot)
                                arr_bot.append(1)
                                j += 1
                            elif arr_pl[j] == 1 and arr_pl[j-1] == 0:
                                rez_bot = rez_bot + 3
                                await true_false(rez_pl, rez_bot)
                                arr_bot.append(0)
                                j += 1
                            elif arr_pl[j] == 0 and arr_pl[j-1] == 1:
                                rez_pl = rez_pl + 3
                                await false_true(rez_pl, rez_bot)
                                arr_bot.append(1)
                                j += 1
                            elif arr_pl[j] == 0 and arr_pl[j-1] == 0:
                                await false_false(rez_pl, rez_bot)
                                arr_bot.append(0)
                                j += 1
                        else:  # Выбрана стртегия Обманщик
                            if arr_pl[j] == 1:
                                rez_bot = rez_bot + 3
                                await true_false(rez_pl, rez_bot)
                                arr_bot.append(0)
                                j += 1
                            else:
                                await false_false(rez_pl, rez_bot)
                                arr_bot.append(0)
                                j += 1
            case 6:  # Обезьяна (случайный выбор хода)
                arr_bot.append(int(random.randint(0, 1)))
                if arr_pl[j] == 0 and arr_bot[j] == 0:
                    await false_false(rez_pl, rez_bot)
                    j += 1
                elif arr_pl[j] == 1 and arr_bot[j] == 1:
                    rez_pl = rez_pl + 2
                    rez_bot = rez_bot + 2
                    await true_true(rez_pl, rez_bot)
                    j += 1
                elif arr_pl[j] == 1 and arr_bot[j] == 0:
                    rez_bot = rez_bot + 3
                    await true_false(rez_pl, rez_bot)
                    j += 1
                elif arr_pl[j] == 0 and arr_bot[j] == 1:
                    rez_pl = rez_pl + 3
                    await false_true(rez_pl, rez_bot)
                    j += 1
            # Подражатель_2 (начинает с доверия, обманывает после 2-х обманов противником)
            case 7:
                if j < 2:
                    if arr_pl[j] == 1:
                        rez_pl = rez_pl + 2
                        rez_bot = rez_bot + 2
                        await true_true(rez_pl, rez_bot)
                        arr_bot.append(1)
                        j += 1
                    else:
                        rez_pl = rez_pl + 3
                        await false_true(rez_pl, rez_bot)
                        arr_bot.append(1)
                        j += 1
                else:
                    if (arr_pl[j] == 1 and arr_pl[j-1] == 1 and arr_pl[j-2] == 1) or (
                            arr_pl[j] == 1 and arr_pl[j-1] == 1 and arr_pl[j-2] == 0) or (
                                arr_pl[j] == 1 and arr_pl[j-1] == 0 and arr_pl[j-2] == 1):
                        rez_pl = rez_pl + 2
                        rez_bot = rez_bot + 2
                        await true_true(rez_pl, rez_bot)
                        arr_bot.append(1)
                        j += 1
                    elif (arr_pl[j] == 0 and arr_pl[j-1] == 1 and arr_pl[j-2] == 1) or (
                            arr_pl[j] == 0 and arr_pl[j-1] == 0 and arr_pl[j-2] == 1) or (
                                arr_pl[j] == 0 and arr_pl[j-1] == 1 and arr_pl[j-2] == 0):
                        rez_pl = rez_pl + 3
                        await false_true(rez_pl, rez_bot)
                        arr_bot.append(1)
                        j += 1
                    elif arr_pl[j] == 1 and arr_pl[j-1] == 0 and arr_pl[j-2] == 0:
                        rez_bot = rez_bot + 3
                        await true_false(rez_pl, rez_bot)
                        arr_bot.append(0)
                        j += 1
                    elif arr_pl[j] == 0 and arr_pl[j-1] == 0 and arr_pl[j-2] == 0:
                        await false_false(rez_pl, rez_bot)
                        arr_bot.append(0)
                        j += 1
            case 8:  # Простак (начинает с доверия, далее повторяет свой предыдущий ход, при обмане меняеет)
                if j == 0:
                    if arr_pl[0] == 1:
                        rez_pl = rez_pl + 2
                        rez_bot = rez_bot + 2
                        await true_true(rez_pl, rez_bot)
                        arr_bot.append(1)
                        j += 1
                    else:
                        rez_pl = rez_pl + 3
                        await false_true(rez_pl, rez_bot)
                        arr_bot.append(1)
                        j += 1
                else:
                    if arr_pl[j] == 1:
                        if (arr_bot[j-1] == 1 and arr_pl[j-1] == 1) or (arr_bot[j-1] == 0 and arr_pl[j-1] == 0):
                            rez_pl = rez_pl + 2
                            rez_bot = rez_bot + 2
                            await true_true(rez_pl, rez_bot)
                            arr_bot.append(1)
                            j += 1
                        else: # (arr_bot[j-1] == 1 and arr_pl[j-1] == 0) or (arr_bot[j-1] == 0 and arr_pl[j-1] == 1):
                            rez_bot = rez_bot + 3
                            await true_false(rez_pl, rez_bot)
                            arr_bot.append(0)
                            j += 1
                    else: # arr_pl[j] == 0:
                        if (arr_bot[j-1] == 1 and arr_pl[j-1] == 1) or (arr_bot[j-1] == 0 and arr_pl[j-1] == 0):
                            rez_pl = rez_pl + 3
                            await false_true(rez_pl, rez_bot)
                            arr_bot.append(1)
                            j += 1
                        else:  # (arr_bot[j-1] == 1 and arr_pl[j-1] == 0) or (arr_bot[j-1] == 0 and arr_pl[j-1] == 1):
                            await false_false(rez_pl, rez_bot)
                            arr_bot.append(0)
                            j += 1
        if j < 10:
            await bot.send_message(chat_id=message.chat.id, text=('Сделайте следующий ход'))
        else:
            logging.info(
                f'{opp} 1-Наивный, 2-Обманщик, 3-Подражатель, 4-Злопамятный, 5-Детектив, 6-Обезьяна, 7-Подражатель_2, 8-Простой')
            logging.info(
                f'Игрок\t\t{arr_pl}, очки игрока {rez_pl}')
            logging.info(
                f'Соперник\t{arr_bot}, очки соперника {rez_bot}')
            if rez_pl > rez_bot:
                await bot.send_message(chat_id=message.chat.id, text=('Поздравляем, вы выиграли'))
            if rez_pl < rez_bot:
                await bot.send_message(chat_id=message.chat.id, text=('Сожалеем Вы проиграли.'))
            if rez_pl == rez_bot:
                await bot.send_message(chat_id=message.chat.id, text=('"Победила Дружба"'))
            await bot.send_message(chat_id=message.chat.id, text='Для продолжения игры, просто сделай ход')
            j, rez_pl, rez_bot = 0, 0, 0
            arr_pl.clear()
            arr_bot.clear()

    async def false_false(x, y):
        await bot.send_message(chat_id=message.chat.id, text='Мы не доверяем друг другу, ни кто из нас не получает награды.')
        await bot.send_message(chat_id=message.chat.id, text=f'Cчет составляет: Игрок {x} монет, Бот {y} монет')

    async def true_false(x, y):
        await bot.send_message(chat_id=message.chat.id, text='Сожалею, я Вас обманул. Вы не получаете награды, я получаю в награду 3 монеты.')
        await bot.send_message(chat_id=message.chat.id, text=f'Cчет составляет: Игрок {x} монет, Бот {y} монет')

    async def false_true(x, y):
        await bot.send_message(chat_id=message.chat.id, text='Вам удалось обмануть меня. Вы получаете в награду 3 монеты, я ничего не получаю.')
        await bot.send_message(chat_id=message.chat.id, text=f'Cчет составляет: Игрок {x} монет, Бот {y} монет')

    async def true_true(x, y):
        await bot.send_message(chat_id=message.chat.id, text='Хорошо, мы доверяем друг другу. Каждый из нас получает в награду по 2 монеты.')
        await bot.send_message(chat_id=message.chat.id, text=f'Счет составляет: Игрок {x} монет, Бот {y} монет')


print('Server  starting')


async def main():
    await dp.start_polling(bot)

if __name__ == '__main__':
    asyncio.run(main())
